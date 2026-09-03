from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def create_sample(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Profiles"
    headers = ["Register No", "Name", "CodeChef", "LeetCode", "HackerRank", "Codeforces", "GFG", "LinkedIn", "GitHub"]
    sheet.append(headers)
    sheet.append([
        "REG-001",
        "Anafa Thabassum S",
        "https://www.codechef.com/users/anafa_sadiq",
        "https://leetcode.com/u/anafasadiq/",
        "https://www.hackerrank.com/profile/anafathabassum",
        None,
        None,
        None,
        "https://github.com/Anafa-thabassum",
    ])
    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor="14213D")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(vertical="center")
    for cell in sheet[2]:
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        if isinstance(cell.value, str) and cell.value.startswith("http"):
            cell.hyperlink = cell.value
            cell.style = "Hyperlink"
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.sheet_view.showGridLines = False
    for index in range(1, len(headers) + 1):
        sheet.column_dimensions[get_column_letter(index)].width = 24 if index > 2 else 20
    sheet.row_dimensions[1].height = 28
    sheet.row_dimensions[2].height = 48
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A3
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 1
    workbook.save(path)


if __name__ == "__main__":
    create_sample(Path(__file__).resolve().parent.parent / "sample_input.xlsx")
