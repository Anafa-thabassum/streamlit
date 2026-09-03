from openpyxl import load_workbook

from tracker.database import Database
from tracker.models import FetchError, ProfileResult
from tracker.reporting import ExcelReportWriter, SUMMARY_HEADERS, verify_report, weekly_progress_rows, weekly_progress_wide_rows


def test_report_has_all_sheets_headers_and_na(tmp_path):
    db = Database(tmp_path / "profiles.db")
    profile_id, _ = db.upsert_profile("Ana", "LeetCode", "ana", "https://leetcode.com/u/ana", "R001")
    db.save_success(profile_id, ProfileResult(
        platform="LeetCode", username="ana", profile_url="https://leetcode.com/u/ana",
        problems_solved=100, easy=50, medium=40, hard=10, rating=1700,
    ))
    path = ExcelReportWriter(db, tmp_path).generate(tmp_path / "report.xlsx")
    assert verify_report(path)["valid"]
    wb = load_workbook(path)
    assert [cell.value for cell in wb["Summary"][1]] == SUMMARY_HEADERS
    assert wb["Summary"]["A2"].value == "R001"
    assert wb["Summary"]["B2"].value == "Ana"
    assert wb["Summary"]["C2"].value == "N/A"
    assert wb["LeetCode"].freeze_panes == "A2"


def test_report_errors_sheet_omits_old_failures_after_success(tmp_path):
    db = Database(tmp_path / "profiles.db")
    profile_id, _ = db.upsert_profile("Ana", "CodeChef", "ana", "https://www.codechef.com/users/ana", "R001")
    db.save_failure(profile_id, FetchError("Rate limited", "CodeChef rate limit reached"))
    db.save_success(profile_id, ProfileResult(
        platform="CodeChef", username="ana", profile_url="https://www.codechef.com/users/ana",
        problems_solved=10, rating=1500,
    ))

    path = ExcelReportWriter(db, tmp_path).generate(tmp_path / "report.xlsx")
    wb = load_workbook(path)
    assert wb["Errors"].max_row == 1


def test_weekly_progress_compares_two_latest_snapshots(tmp_path):
    db = Database(tmp_path / "profiles.db")
    profile_id, _ = db.upsert_profile("Ana", "Codeforces", "ana", "https://codeforces.com/profile/ana")
    db.save_success(profile_id, ProfileResult(platform="Codeforces", username="ana", profile_url="x", problems_solved=10, rating=1200))
    db.save_success(profile_id, ProfileResult(platform="Codeforces", username="ana", profile_url="x", problems_solved=14, rating=1255))
    row = weekly_progress_rows(db)[0]
    assert row["Problems Added"] == 4
    assert row["Rating Change"] == 55


def test_weekly_progress_wide_rows_keeps_student_platforms_on_one_row():
    headers, rows = weekly_progress_wide_rows([
        {
            "Register No": "310624148101",
            "Name": "A Smirrithi",
            "Platform": "CodeChef",
            "Problems Solved Before": 259,
            "Problems Solved Now": 260,
            "Problems Added": 1,
            "Rating Before": 1829,
            "Rating Now": 1835,
            "Rating Change": 6,
            "Rank Before": 3375,
            "Rank Now": 3200,
            "Rank Change": -175,
            "Fetch Date": "2026-09-03",
        },
        {
            "Register No": "310624148101",
            "Name": "A Smirrithi",
            "Platform": "LeetCode",
            "Problems Solved Before": 609,
            "Problems Solved Now": 611,
            "Problems Added": 2,
            "Rating Before": None,
            "Rating Now": None,
            "Rating Change": None,
            "Rank Before": 136376,
            "Rank Now": 135000,
            "Rank Change": -1376,
            "Fetch Date": "2026-09-03",
        },
    ])

    assert len(rows) == 1
    assert rows[0]["Register No"] == "310624148101"
    assert rows[0]["CodeChef Problems Solved Now"] == 260
    assert rows[0]["LeetCode Problems Solved Now"] == 611
    assert "Platform" not in headers
