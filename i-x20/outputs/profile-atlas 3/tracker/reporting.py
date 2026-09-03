from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .database import Database

NA = "N/A"
PLATFORM_SHEETS = ["CodeChef", "LeetCode", "HackerRank", "Codeforces", "GFG", "LinkedIn", "GitHub"]
WEEKLY_PROGRESS_METRICS = [
    "Problems Solved Before", "Problems Solved Now", "Problems Added",
    "Rating Before", "Rating Now", "Rating Change",
    "Rank Before", "Rank Now", "Rank Change", "Fetch Date",
]
SUMMARY_HEADERS = [
    "Register No", "Name", "CodeChef Rating", "CodeChef Rank", "CodeChef Problems Solved",
    "CodeChef Total Submissions", "CodeChef Contests",
    "LeetCode Rating", "LeetCode Global Rank", "LeetCode Problems Solved",
    "LeetCode Easy", "LeetCode Medium", "LeetCode Hard", "Codeforces Rating",
    "Codeforces Rank", "Codeforces Problems Solved", "Codeforces Total Submissions", "Codeforces Contests",
    "HackerRank Rank", "HackerRank Hackos",
    "HackerRank Problems Solved", "GFG Rank", "GFG Coding Score", "GFG Problems Solved",
    "GFG Easy", "GFG Medium", "GFG Hard",
    "GitHub Repositories", "GitHub Followers", "GitHub Contributions",
    "LinkedIn Connections/Followers", "Last Updated",
]

COMMON_DETAIL_COLUMNS = [
    ("Register No", "student_register_no"), ("Name", "student_name"),
    ("Username", "username"), ("Profile URL", "profile_url"),
    ("Fetch Status", "fetch_status"), ("Last Updated", "last_fetched"),
]
PLATFORM_DETAIL_COLUMNS = {
    "CodeChef": COMMON_DETAIL_COLUMNS + [
        ("Current Rating", "rating"), ("Highest Rating", "max_rating"), ("Global Rank", "global_rank"),
        ("Country Rank", "country_rank"), ("Stars", "stars"), ("Problems Solved", "problems_solved"),
        ("Total Submissions", "total_submissions"), ("Contests Attended", "contests_attended"),
        ("Last Activity", "last_activity"),
    ],
    "LeetCode": COMMON_DETAIL_COLUMNS + [
        ("Global Rank", "global_rank"), ("Problems Solved", "problems_solved"), ("Easy", "easy"),
        ("Medium", "medium"), ("Hard", "hard"), ("Contest Rating", "contest_rating"),
        ("Contest Rank", "contest_rank"), ("Contests Attended", "contests_attended"),
        ("Total Submissions", "total_submissions"), ("Acceptance Rate (%)", "acceptance_rate"),
        ("Reputation", "reputation"), ("Badges", "badges_json"), ("Last Activity", "last_activity"),
    ],
    "HackerRank": COMMON_DETAIL_COLUMNS + [
        ("Rank", "rank"), ("Hackos", "hackos"), ("Problems Solved", "problems_solved"), ("Badges", "badges_json"),
        ("Certificates", "certificates_json"), ("Followers", "followers"), ("Last Activity", "last_activity"),
    ],
    "Codeforces": COMMON_DETAIL_COLUMNS + [
        ("Rating", "rating"), ("Max Rating", "max_rating"), ("Rank", "rank"), ("Max Rank", "max_rank"),
        ("Problems Solved", "problems_solved"), ("Contests Attended", "contests_attended"),
        ("Total Submissions", "total_submissions"), ("Display Name", "name"), ("Last Activity", "last_activity"),
    ],
    "GFG": COMMON_DETAIL_COLUMNS + [
        ("Coding Score", "coding_score"), ("Rank", "rank"), ("Problems Solved", "problems_solved"),
        ("Easy", "easy"), ("Medium", "medium"), ("Hard", "hard"),
        ("Articles", "articles"), ("Last Activity", "last_activity"),
    ],
    "LinkedIn": COMMON_DETAIL_COLUMNS + [
        ("Display Name", "name"), ("Headline", "headline"), ("Followers", "followers"),
        ("Connections", "connections"),
    ],
    "GitHub": COMMON_DETAIL_COLUMNS + [
        ("Display Name", "name"), ("Bio", "bio"), ("Followers", "followers"), ("Following", "following_count"),
        ("Repositories", "repositories"), ("Stars Received", "stars_received"), ("Forks", "forks"),
        ("Contributions", "contributions"), ("Account Created", "account_created"), ("Last Activity", "last_activity"),
    ],
}


def _value(value: Any) -> Any:
    return NA if value is None or value == "" else value


def latest_snapshot_rows(database: Database, *, active_only: bool = True) -> list[dict[str, Any]]:
    active_clause = "WHERE p.active=1" if active_only else ""
    return database.query(f"""
        WITH ranked AS (
          SELECT s.*, ROW_NUMBER() OVER (PARTITION BY s.profile_id ORDER BY s.fetched_at DESC, s.id DESC) AS rn
          FROM profile_snapshots s
        )
        SELECT p.id AS profile_id, p.student_register_no, p.student_name, p.platform, p.username, p.profile_url,
               p.fetch_status, p.last_fetched, p.last_error, r.*
        FROM profiles p LEFT JOIN ranked r ON r.profile_id=p.id AND r.rn=1
        {active_clause} ORDER BY p.student_name COLLATE NOCASE, p.platform
    """)


def weekly_progress_rows(database: Database) -> list[dict[str, Any]]:
    return database.query("""
        WITH ranked AS (
          SELECT s.*, ROW_NUMBER() OVER (PARTITION BY s.profile_id ORDER BY s.fetched_at DESC, s.id DESC) AS rn
          FROM profile_snapshots s
        )
        SELECT p.student_register_no AS "Register No", p.student_name AS Name, p.platform AS Platform,
          old.problems_solved AS "Problems Solved Before", new.problems_solved AS "Problems Solved Now",
          CASE WHEN old.problems_solved IS NOT NULL AND new.problems_solved IS NOT NULL THEN new.problems_solved-old.problems_solved END AS "Problems Added",
          old.rating AS "Rating Before", new.rating AS "Rating Now",
          CASE WHEN old.rating IS NOT NULL AND new.rating IS NOT NULL THEN new.rating-old.rating END AS "Rating Change",
          COALESCE(old.global_rank, CASE WHEN old.rank NOT GLOB '*[^0-9]*' THEN CAST(old.rank AS INTEGER) END) AS "Rank Before",
          COALESCE(new.global_rank, CASE WHEN new.rank NOT GLOB '*[^0-9]*' THEN CAST(new.rank AS INTEGER) END) AS "Rank Now",
          CASE
            WHEN COALESCE(old.global_rank, CASE WHEN old.rank NOT GLOB '*[^0-9]*' THEN CAST(old.rank AS INTEGER) END) IS NOT NULL
             AND COALESCE(new.global_rank, CASE WHEN new.rank NOT GLOB '*[^0-9]*' THEN CAST(new.rank AS INTEGER) END) IS NOT NULL
            THEN COALESCE(new.global_rank, CAST(new.rank AS INTEGER))-COALESCE(old.global_rank, CAST(old.rank AS INTEGER))
          END AS "Rank Change",
          new.fetched_at AS "Fetch Date"
        FROM profiles p JOIN ranked new ON new.profile_id=p.id AND new.rn=1
        LEFT JOIN ranked old ON old.profile_id=p.id AND old.rn=2
        WHERE p.active=1
        ORDER BY p.student_name COLLATE NOCASE, p.platform
    """)


def weekly_progress_wide_headers(platforms: Iterable[str]) -> list[str]:
    platform_set = set(platforms)
    ordered_platforms = [platform for platform in PLATFORM_SHEETS if platform in platform_set]
    extra_platforms = sorted(platform_set - set(PLATFORM_SHEETS))
    headers = ["Register No", "Name"]
    for platform in [*ordered_platforms, *extra_platforms]:
        headers.extend(f"{platform} {metric}" for metric in WEEKLY_PROGRESS_METRICS)
    return headers


def weekly_progress_wide_rows(rows: Iterable[dict[str, Any]]) -> tuple[list[str], list[dict[str, Any]]]:
    source_rows = list(rows)
    platforms = [row.get("Platform") for row in source_rows if row.get("Platform")]
    headers = weekly_progress_wide_headers(platforms)
    grouped: dict[tuple[Any, Any], dict[str, Any]] = {}

    for row in source_rows:
        key = (row.get("Register No"), row.get("Name"))
        grouped.setdefault(key, {"Register No": row.get("Register No"), "Name": row.get("Name")})
        platform = row.get("Platform")
        if not platform:
            continue
        for metric in WEEKLY_PROGRESS_METRICS:
            grouped[key][f"{platform} {metric}"] = row.get(metric)

    output = []
    for grouped_row in grouped.values():
        output.append({header: grouped_row.get(header) for header in headers})
    return headers, output


class ExcelReportWriter:
    def __init__(self, database: Database, reports_dir: Path):
        self.database = database
        self.reports_dir = reports_dir

    @staticmethod
    def _style_sheet(sheet, *, conditional_columns: Iterable[int] = ()) -> None:
        navy = "14213D"
        teal = "00A6A6"
        header_fill = PatternFill("solid", fgColor=navy)
        header_font = Font(color="FFFFFF", bold=True)
        thin = Side(style="thin", color="DDE3EA")
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        sheet.sheet_view.showGridLines = False
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        sheet.row_dimensions[1].height = 34
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.border = Border(bottom=thin)
                cell.alignment = Alignment(vertical="top")
                if isinstance(cell.value, str) and cell.value.startswith("http"):
                    cell.hyperlink = cell.value
                    cell.style = "Hyperlink"
        for index, column in enumerate(sheet.columns, start=1):
            values = [str(cell.value) if cell.value is not None else "" for cell in list(column)[:250]]
            width = min(44, max(11, max((len(value) for value in values), default=0) + 2))
            sheet.column_dimensions[get_column_letter(index)].width = width
        for column in conditional_columns:
            if sheet.max_row > 1:
                letter = get_column_letter(column)
                sheet.conditional_formatting.add(
                    f"{letter}2:{letter}{sheet.max_row}",
                    ColorScaleRule(start_type="min", start_color="FEE2E2", mid_type="percentile", mid_value=50,
                                   mid_color="FEF3C7", end_type="max", end_color="CCFBF1"),
                )
        sheet.sheet_properties.tabColor = teal
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.orientation = "landscape"
        sheet.page_setup.paperSize = sheet.PAPERSIZE_A3
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0
        sheet.print_options.horizontalCentered = True

    @staticmethod
    def _append(sheet, headers: list[str], rows: Iterable[Iterable[Any]]) -> None:
        sheet.append(headers)
        for row in rows:
            values = []
            for header, value in zip(headers, row):
                if isinstance(value, str) and any(word in header.lower() for word in ("date", "updated", "activity", "created")):
                    try:
                        value = datetime.fromisoformat(value.replace("Z", "+00:00")).replace(tzinfo=None)
                    except ValueError:
                        pass
                values.append(_value(value))
            sheet.append(values)
        for row in sheet.iter_rows(min_row=2):
            for cell, header in zip(row, headers):
                if isinstance(cell.value, datetime):
                    cell.number_format = "yyyy-mm-dd hh:mm"

    def _summary(self, rows: list[dict[str, Any]]) -> list[list[Any]]:
        students: dict[tuple[str, str], dict[str, dict[str, Any]]] = {}
        for row in rows:
            key = (row.get("student_register_no") or "", row["student_name"])
            students.setdefault(key, {})[row["platform"]] = row
        output = []
        for (register_no, name), platforms in students.items():
            cc, lc, cf = platforms.get("CodeChef", {}), platforms.get("LeetCode", {}), platforms.get("Codeforces", {})
            hr, gfg, gh, li = (platforms.get("HackerRank", {}), platforms.get("GFG", {}),
                                platforms.get("GitHub", {}), platforms.get("LinkedIn", {}))
            last = max((p.get("last_fetched") for p in platforms.values() if p.get("last_fetched")), default=None)
            output.append([
                register_no or None, name, cc.get("rating"), cc.get("global_rank") or cc.get("rank"), cc.get("problems_solved"),
                cc.get("total_submissions"), cc.get("contests_attended"),
                lc.get("rating"), lc.get("global_rank"), lc.get("problems_solved"), lc.get("easy"), lc.get("medium"), lc.get("hard"),
                cf.get("rating"), cf.get("rank"), cf.get("problems_solved"), cf.get("total_submissions"), cf.get("contests_attended"),
                hr.get("rank"), hr.get("hackos"), hr.get("problems_solved"),
                gfg.get("rank"), gfg.get("coding_score"), gfg.get("problems_solved"), gfg.get("easy"), gfg.get("medium"), gfg.get("hard"),
                gh.get("repositories"), gh.get("followers"), gh.get("contributions"),
                li.get("connections") or li.get("followers"), last,
            ])
        return output

    def generate(self, path: Path | None = None) -> Path:
        self.reports_dir.mkdir(parents=True, exist_ok=True)
        if path is None:
            base = self.reports_dir / f"profile_report_{date.today().isoformat()}.xlsx"
            path = base
            serial = 2
            while path.exists():
                path = self.reports_dir / f"profile_report_{date.today().isoformat()}_{serial}.xlsx"
                serial += 1
        wb = Workbook()
        wb.remove(wb.active)
        latest = latest_snapshot_rows(self.database)

        summary = wb.create_sheet("Summary")
        self._append(summary, SUMMARY_HEADERS, self._summary(latest))
        self._style_sheet(summary, conditional_columns=(3, 6, 12, 19, 20))

        progress_headers, progress_rows = weekly_progress_wide_rows(weekly_progress_rows(self.database))
        progress = wb.create_sheet("Weekly Progress")
        self._append(progress, progress_headers, ([row.get(h) for h in progress_headers] for row in progress_rows))
        self._style_sheet(progress)

        for platform in PLATFORM_SHEETS:
            sheet = wb.create_sheet(platform)
            columns = PLATFORM_DETAIL_COLUMNS[platform]
            headers = [label for label, _ in columns]
            detail = []
            for row in latest:
                if row["platform"] != platform:
                    continue
                values = []
                for _, key in columns:
                    value = row.get(key)
                    if key in {"badges_json", "certificates_json"}:
                        value = ", ".join(__import__("json").loads(value or "[]"))
                    values.append(value)
                detail.append(values)
            self._append(sheet, headers, detail)
            conditional = tuple(i for i, (_, key) in enumerate(columns, 1) if key in {"rating", "problems_solved", "followers", "repositories"})
            self._style_sheet(sheet, conditional_columns=conditional)

        activities = self.database.query("""
            SELECT p.student_register_no AS "Register No", p.student_name AS Name,
              p.platform AS Platform, p.username AS Username,
              a.activity_date AS Date, a.activity_type AS "Activity Type", a.title AS "Problem/Contest/Repository",
              a.difficulty AS Difficulty, a.status AS Status, a.rating_change AS "Rating Change", a.url AS URL
            FROM activities a JOIN profiles p ON p.id=a.profile_id
            ORDER BY COALESCE(a.activity_date, a.first_seen_at) DESC
        """)
        activity_headers = ["Register No", "Name", "Platform", "Username", "Date", "Activity Type", "Problem/Contest/Repository", "Difficulty", "Status", "Rating Change", "URL"]
        activity_sheet = wb.create_sheet("Activity Log")
        self._append(activity_sheet, activity_headers, ([row.get(h) for h in activity_headers] for row in activities))
        self._style_sheet(activity_sheet, conditional_columns=(10,))

        errors = self.database.query("""
            SELECT "Register No", Name, Platform, "Profile URL", "Error Type", "Error Message", Timestamp FROM (
              SELECT p.student_register_no AS "Register No", p.student_name AS Name,
                p.platform AS Platform, p.profile_url AS "Profile URL",
                latest.error_type AS "Error Type", latest.error_message AS "Error Message", latest.fetched_at AS Timestamp
              FROM profiles p JOIN (
                SELECT f.*, ROW_NUMBER() OVER (PARTITION BY f.profile_id ORDER BY f.fetched_at DESC, f.id DESC) rn
                FROM fetch_logs f
              ) latest ON latest.profile_id=p.id AND latest.rn=1
              WHERE p.fetch_status='Failed' AND latest.status='Failed'
              UNION ALL
              SELECT NULL, student_name, platform, profile_url, error_type, error_message, created_at
              FROM import_errors
            ) ORDER BY Timestamp DESC
        """)
        error_headers = ["Register No", "Name", "Platform", "Profile URL", "Error Type", "Error Message", "Timestamp"]
        error_sheet = wb.create_sheet("Errors")
        self._append(error_sheet, error_headers, ([row.get(h) for h in error_headers] for row in errors))
        self._style_sheet(error_sheet)

        wb.properties.title = "Competitive Programming & Developer Profile Report"
        wb.properties.subject = "Public profile statistics and weekly progress"
        wb.properties.creator = "Profile Tracker"
        wb.save(path)
        self.database.record_report(path)
        return path


def verify_report(path: Path) -> dict[str, Any]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=False, data_only=False)
    expected = ["Summary", "Weekly Progress", *PLATFORM_SHEETS, "Activity Log", "Errors"]
    missing = [sheet for sheet in expected if sheet not in wb.sheetnames]
    return {
        "valid": not missing and wb.sheetnames[:2] == ["Summary", "Weekly Progress"],
        "missing_sheets": missing,
        "sheet_count": len(wb.sheetnames),
        "sheets": wb.sheetnames,
    }
